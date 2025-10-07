import openpyxl
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

import WebDriverInter as InterW

# --- CONFIGURACIÓN ---
EXCEL_FILE = 'EstablecerRuta.xlsx'
SHEET_NAME = 'Hoja1'
GUIA_COLUMN = 1  # Columna A
RESULTS_START_COLUMN = 2  # Columna B


# --- MEJORAS APLICADAS ---
# 1. Manejo de errores con try/finally para asegurar que el navegador se cierre y el excel se guarde.
# 2. El guardado del archivo Excel se mueve fuera del bucle para mayor eficiencia.
# 3. Se añade un try/except dentro del bucle para que un error en una guía no detenga todo el proceso.


def main():
    # driver = None
    try:
        # --- 1. CONECTARSE AL EXCEL Y LEER DATOS ---
        print("Cargando archivo Excel...")
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME]
        print(f"Archivo '{EXCEL_FILE}' cargado. Hoja '{SHEET_NAME}' seleccionada.")
        print(f"Total de filas a procesar: {ws.max_row - 1}")  # Restamos 1 para no contar el encabezado

        # --- 2. INICIAR NAVEGADOR Y LOGIN ---
        print("Iniciando navegador Chrome...")
        service = ChromeService(executable_path=ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service)
        driver.get("https://www3.interrapidisimo.com/SitioLogin/auth/login")

        print("Realizando login en la web...")
        InterW.login(driver)
        InterW.SelectApliEnvio(driver)

        # --- 3. BUCLE PRINCIPAL DE PROCESAMIENTO ---
        # Empezamos desde la fila 2  saltar el encabezado si existe.
        for row_num in range(2, ws.max_row + 1):
            try:
                # Leer el número de guía directamente de la celda actual
                guia = ws.cell(row=row_num, column=GUIA_COLUMN).value

                # Si la celda de la guía está vacía, saltamos a la siguiente fila
                if not guia:
                    print(f"Fila {row_num}: La celda de la guía está vacía. Saltando...")
                    continue

                print("-" * 50)
                print(f"Procesando Fila {row_num} | Guía: {guia}")

                # CONDICIÓN procesar solo si la columna 6 está vacía
                if ws.cell(row=row_num, column=6).value is not None:
                     print(f"Fila {row_num}: La guía '{guia}' ya parece haber sido procesada (Columna 6 no está vacía). Saltando...")
                     continue

                print(f"Buscando información para la guía {guia}...")
                VerificaGuia, Telefono, CiudadDestino, TipoEntrega, Direccion = InterW.BuscarGuia(driver, guia)

                # Escribir los resultados en el Excel
                ws.cell(row=row_num, column=RESULTS_START_COLUMN, value=VerificaGuia)
                ws.cell(row=row_num, column=RESULTS_START_COLUMN + 1, value=Telefono)
                ws.cell(row=row_num, column=RESULTS_START_COLUMN + 2, value=Direccion)
                ws.cell(row=row_num, column=RESULTS_START_COLUMN + 3, value=TipoEntrega)
                ws.cell(row=row_num, column=RESULTS_START_COLUMN + 4, value=CiudadDestino)  # Columna 6

                print(f"Fila {row_num}: Datos guardados exitosamente.")

            except Exception as e:
                print(f"ERROR en la fila {row_num} con la guía '{guia}': {e}")
                print("Continuando con la siguiente guía...")
                # Opcional: puedes escribir el error en una columna del Excel
                ws.cell(row=row_num, column=7, value=f"ERROR: {e}")

    except Exception as e:
        print(f"Ha ocurrido un error crítico en el programa: {e}")
    finally:
        # --- 4. LIMPIEZA FINAL ---
        if driver:
            print("Cerrando el navegador...")
            driver.quit()

        print("Guardando los cambios en el archivo Excel...")
        wb.save(EXCEL_FILE)
        print("Proceso finalizado.")


if __name__ == "__main__":
    main()