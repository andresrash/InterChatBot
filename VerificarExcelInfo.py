import time

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

import WebDriverInter as InterW

# conectarse al excel
wb = openpyxl.load_workbook('EstablecerRuta.xlsx')
ws = wb.active
ws = wb['Hoja1']
print('Total numero de filas : '+str(ws.max_row)+'. y total de columnas: '+str(ws.max_column))

# obtener los datos de la 1 fila
values = [ws.cell(row=i,column=1).value for i in range(1,ws.max_row+1)]
print(values)

service = ChromeService(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)
driver.get("https://www3.interrapidisimo.com/SitioLogin/auth/login")

InterW.login(driver)
InterW.SelectApliEnvio(driver)

for i in range(1,ws.max_row+1):
    print("imprimir valor:", values[i-1])
    print("for i =", i, " Url:", driver.current_url)
    print("Guia =", values[i-1], " Url:", driver.current_url)

    time.sleep(5)
    driver.implicitly_wait(10)
    VerificaGuia, Telefono, CiudadDestino, TipoEntrega, Direccion = InterW.BuscarGuia(driver,values[i-1])
    driver.implicitly_wait(10)
    time.sleep(5)

    wb.save('EstablecerRuta.xlsx')
    ws.cell(row=i,column=2, value=VerificaGuia)
    ws.cell(row=i,column=3, value=Telefono)
    ws.cell(row=i,column=4, value=Direccion)
    ws.cell(row=i,column=5, value=TipoEntrega)
    ws.cell(row=i,column=6, value=CiudadDestino)
    wb.save('EstablecerRuta.xlsx')

