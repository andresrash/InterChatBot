import time
import requests
import openpyxl
import pywhatkit as kit

# Datos necesarios

contact_no = "+573137048612"
name = "Cristian Cardona"
image_path = "images/greetings.jpg"
message = "Hola " + name
wait_time = 10 # tiempo que espera antes de enviar el mensaje en minutos
close_tab = True # Cerrar la pestaña del navegador 
close_time = 20 # tiempo para cerrar la pestaña en minutos


# conectarse al excel 
wb = openpyxl.load_workbook('EnvioMensajes.xlsx')
ws = wb.active
ws = wb['Hoja1']
print('Total numero de filas : '+str(ws.max_row)+'. y total de columnas: '+str(ws.max_column))

# obtener los datos de la 1 fila
values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
print(values)

# obtener la matriz
values = [
    [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    for r in range(1,ws.max_row+1)  # filas 1 al maximo de filas 
]
print(values)


# libreria que envia automaticamente los mensajes
# sendwhatmsg_instantly(phone_no: str, message: str, wait_time: int = 15, tab_close: bool = False, close_time: int = 3) -> None

#kit.sendwhatmsg_instantly(contact_no, message, wait_time, close_tab)
time.sleep(5)

# prueba envio de un mensaje con plantilla Hello_word al telefono 3137048612 de prueba en Curl 
# bearer es la clave acceso de 60 dias / 09-10-2025

"""
curl -i -X POST `
  https://graph.facebook.com/v22.0/794047857130311/messages `
  -H 'Authorization: Bearer E1AAK3qHULRIUBPviZBClInHE1t1Mm8bI7ozZAKGrCM6fZCLjrSlcp5jeWX5c79Xug4TuPJ6YANG2YjYMEzpuxn6m1YIkCkVHvkqtNYAwwaYIwRBvosFJMPQYXAusoB4FVWeW8Bdu4i75e08Nh0JgHdLVF30ijLxIavuEcruskYuY3nPE0u4bt7K7hXAzZBtNgWzN8XkIqrQ4P6vBPBZCzaR9X9w7PX5yYb3csMbGtGeW0ZD' `
  -H 'Content-Type: application/json' `
  -d '{ \"messaging_product\": \"whatsapp\", \"to\": \"573137048612\", \"type\": \"template\", \"template\": { \"name\": \"hello_world\", \"language\": { \"code\": \"en_US\" } } }'
"""

# prueba con pyton 

url = "https://graph.facebook.com/v22.0/794047857130311/messages"

headers = {
    "Authorization": "Bearer EAAK3qHULRIUBPviZBClInHE1t1Mm8bI7ozZAKGrCM6fZCLjrSlcp5jeWX5c79Xug4TuPJ6YANG2YjYMEzpuxn6m1YIkCkVHvkqtNYAwwaYIwRBvosFJMPQYXAusoB4FVWeW8Bdu4i75e08Nh0JgHdLVF30ijLxIavuEcruskYuY3nPE0u4bt7K7hXAzZBtNgWzN8XkIqrQ4P6vBPBZCzaR9X9w7PX5yYb3csMbGtGeW0ZD",
    "Content-Type": "application/json"
}

payload = {
    "messaging_product": "whatsapp",
    "to": "573137048612",
    "type": "template",
    "template": {
        "name": "hello_world",
        "language": {
            "code": "en_US"
        }
    }
}

response = requests.post(url, headers=headers, json=payload)

print("Status Code:", response.status_code)
print("Response:", response.json())

